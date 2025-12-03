import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

class StatsLogger:
    EXCEL_FILE = "Experiments_History.xlsx"
    STANDARD_COLUMNS = [
        "Timestamp",
        "Algorithm",
        "Target Word",
        "Word Length",
        "Time (s)",
        "Expanded Nodes",
        "Total Guesses",
        "Solution Path",
        "Max Memory (nodes)",
        "Status"
    ]
    @staticmethod
    def save_run(algorithm_name, stats_dict, solution_path, target_word, word_length=5):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        time_taken = stats_dict.get("Time", stats_dict.get("search_time_sec", "N/A"))
        expanded = stats_dict.get("Expanded Nodes", stats_dict.get("expanded_nodes", 0))
        total_guesses = stats_dict.get("Total Guesses", stats_dict.get("Guesses", stats_dict.get("total_guesses", len(solution_path))))
        max_memory = stats_dict.get("Max Queue", stats_dict.get("Max Queue Size", stats_dict.get("max_memory_nodes", "N/A")))
        status = stats_dict.get("Result", stats_dict.get("Status", "Unknown"))
        path_str = " → ".join(solution_path[:15])  
        if len(solution_path) > 15:
            path_str += f" ... (+{len(solution_path) - 15} more)"
        row_data = [
            timestamp,
            algorithm_name,
            target_word,
            word_length,
            time_taken,
            expanded,
            total_guesses,
            path_str,
            max_memory,
            status
        ]
        StatsLogger._append_to_excel(row_data)
        print(f"\nStatistics saved to {StatsLogger.EXCEL_FILE}")
    
    @staticmethod
    def _append_to_excel(row_data):
        file_path = StatsLogger.EXCEL_FILE
        if os.path.exists(file_path):
            
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Experiments"
            ws.append(StatsLogger.STANDARD_COLUMNS)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            column_widths = [20, 12, 12, 12, 12, 15, 15, 50, 18, 12]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(1, i).column_letter].width = width
        ws.append(row_data)
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        algo_col = 2  
        algo_colors = {
            "BFS": "E2EFDA",
            "DFS": "FCE4D6",
            "A*": "DDEBF7",
            "Entropy": "FFF2CC"
        }
        algo_cell = ws.cell(last_row, algo_col)
        algo_name = algo_cell.value
        if algo_name in algo_colors:
            algo_cell.fill = PatternFill(start_color=algo_colors[algo_name], 
                                        end_color=algo_colors[algo_name], 
                                        fill_type="solid")
        wb.save(file_path)
    @staticmethod
    def print_stats(algorithm_name, stats_dict):
        print("\n" + "="*50)
        print(f"{algorithm_name.upper()} SOLVER STATISTICS")
        print("="*50)
        if "Time" in stats_dict:
            print(f"Time             : {stats_dict['Time']}")
        elif "search_time_sec" in stats_dict:
            print(f"Time             : {stats_dict['search_time_sec']:.4f}s")
        
        if "Expanded Nodes" in stats_dict:
            print(f"Expanded Nodes   : {stats_dict['Expanded Nodes']}")
        elif "expanded_nodes" in stats_dict:
            print(f"Expanded Nodes   : {stats_dict['expanded_nodes']}")
        
        if "Total Guesses" in stats_dict:
            print(f"Total Guesses    : {stats_dict['Total Guesses']}")
        elif "Guesses" in stats_dict:
            print(f"Total Guesses    : {stats_dict['Guesses']}")
        elif "total_guesses" in stats_dict:
            print(f"Total Guesses    : {stats_dict['total_guesses']}")
        
        if "Max Queue" in stats_dict:
            print(f"Max Queue Size   : {stats_dict['Max Queue']}")
        elif "Max Queue Size" in stats_dict:
            print(f"Max Queue Size   : {stats_dict['Max Queue Size']}")
        
        if "Memory Usage" in stats_dict:
            print(f"Memory Usage     : {stats_dict['Memory Usage']}")
        
        if "Winning Steps" in stats_dict:
            print(f"Winning Steps    : {stats_dict['Winning Steps']}")
        
        if "Result" in stats_dict:
            print(f"Status           : {stats_dict['Result']}")
        elif "Status" in stats_dict:
            print(f"Status           : {stats_dict['Status']}")
        
        print("="*50 + "\n")
