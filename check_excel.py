"""Check Excel history"""
from openpyxl import load_workbook

wb = load_workbook('Experiments_History.xlsx')
ws = wb.active

print(f'Total rows: {ws.max_row} (including header)')
print(f'Data rows: {ws.max_row - 1}')
print('\nAll entries:')
for row in ws.iter_rows(min_row=2, values_only=True):
    print(f'  {row[0]} | {row[1]:8s} | {row[2]:6s} | Guesses: {row[6]}')
